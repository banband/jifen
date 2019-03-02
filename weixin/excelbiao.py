import openpyxl



class exel():

    def readExel(self,lujing):
        self.lujing = lujing
        #filename = r'D:\work\Excel_txtProcesss\new-微博-合并\58.xlsx'
        inwb = openpyxl.load_workbook(self.lujing)  # 读文件

        sheetnames = inwb.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
        print(sheetnames)
        ws = inwb.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容
        print(ws)

        # 获取sheet的最大行数和列数
        rows = ws.max_row
        cols = ws.max_column
        for r in range(1, rows):
            for c in range(1, cols):
                print(ws.cell(r, c).value)
            if r == 10:
                break


    def writeExcel(self,row,col,value,lujing):
        self.row = row
        self.col = col
        self.value = value
        self.lujing = lujing
        outwb = openpyxl.Workbook()  # 打开一个将写的文件
        outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
        for row in range(1, 70000):
            for col in range(1, 4):
                outws.cell(self.row, self.col).self.value = row * 2  # 写文件
            print(row)
        #saveExcel = "D:\\work\\Excel_txtProcesss\\test.xlsx"
        outwb.save(self.lujing)  # 一定要记得保存
if __name__ == '__main__':
    jilu = exel()
    #jilu.readExel("jilu.xlsx")
    jilu.writeExcel("jilu.xlsx")